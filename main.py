import asyncio
import datetime
import json
from typing import Optional
import aiohttp
from aiohttp_retry import ExponentialRetry, RetryClient
import requests
from openpyxl import load_workbook
import openpyxl
from fake_useragent import UserAgent


class ParserRivegauche:
    """
    Класс для парсинга данных с сайта Rivegauche.
    """

    def __init__(self):
        """
        Инициализация объекта ParserRivegauche.
        """
        self.ua = UserAgent()
        self.data_file = 'Задание_для_Разработчика_парсинга.xlsx'
        self.url_rivegauche = 'https://rivegauche.ru'
        self.api_url_products = 'https://api.rivegauche.ru/rg/v1/newRG/products/'
        self.api_url_brands = 'https://api.rivegauche.ru/rg/v1/newRG/products/search'
        self.brand_code_payot = 'rg_brand_181'
        self.category_code_condtioner = 'HairCare_Condtioner'
        self.params = {
            'fields': 'FULL',
            'pageSize': '100',
        }
        self.links_products = []
        self.info_products = {}

    def get_links_from_file(self) -> None:
        """
        Извлекает ссылки на продукты из файла и сохраняет их в self.links_products.
        """
        workbook = load_workbook(filename=self.data_file)
        sheet = workbook['Задание 1']
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
            for cell in row:
                if cell.value:
                    self.links_products.append(cell.value.strip())

    async def get_info_about_products(self, session: aiohttp.ClientSession, base_link: str) -> None:
        """
        Асинхронно получает информацию о продуктах с сайта Rivegauche.
        """
        retry_options = ExponentialRetry(attempts=5)  # параметр для повторных попыток подключения
        retry_client = RetryClient(raise_for_status=False, retry_options=retry_options, client_session=session,
                                   start_timeout=0.5)  # контекстный менеджер
        self.info_products.clear()  # очистка словаря
        url = f"{self.api_url_products}{base_link.split('/')[-1]}"
        async with retry_client.get(url=url) as response:
            if response.ok:
                print(base_link)
                resp = await response.text()
                data_json = json.loads(resp)  # полученный json из запроса
                status = data_json['maxQuantity'] != 0  # доступность для заказа(остаток товар больше 0)
                categories_chain = data_json['categoriesChain']  # категории товара
                brand_name = data_json['brand']['name']  # название бренд
                name_product = data_json['name']  # имя товара
                price_discount = int(data_json['prices'][0]['value'])  # цена со скидкой(первая)
                full_price = int(data_json['prices'][-1]['value'])  # цена без скидки
                date_now = datetime.datetime.now().strftime("%d-%m-%y")  # текущая дата
                # запись полученных данных в словарь
                self.info_products[base_link] = {
                    'Наименование': name_product,
                    'Ссылка': base_link,
                    'Цена до скидки': full_price,
                    'Цена со скидкой или по карте лояльности': price_discount,
                    'Доступен для заказа (есть остаток)': status,
                    'Цепочка категорий': categories_chain,
                    'Бренд': brand_name,
                    'Дата': date_now
                }

    async def create_async_tasks(self) -> None:
        """
        Создает и запускает асинхронные задачи.
        """
        connector = aiohttp.TCPConnector(force_close=True)
        async with aiohttp.ClientSession(headers={'User-Agent': self.ua.random}, connector=connector) as session:
            tasks = [asyncio.create_task(self.get_info_about_products(session, base_link)) for base_link in
                     self.links_products]  # создание задач для асинхронного выполнения
            await asyncio.gather(*tasks)

    def get_link_from_products(self, brand_code: Optional[str], category_code: Optional[str]) -> None:
        """
        Получает ссылки на продукты для брендов или категорий.
        """
        self.links_products.clear()
        for page in range(1000):
            self.params['currentPage'] = page
            if brand_code:
                self.params['brandCode'] = brand_code
            if category_code:
                self.params['categoryCode'] = category_code
            # отправка запроса с необходимо сформированными параметрами
            response = requests.get(self.api_url_brands, params=self.params, headers={'User-Agent': self.ua.random})
            data_json = json.loads(response.text)
            data_result = data_json['results']
            if data_result:
                for result in data_json['results']:
                    result_url = f"{self.url_rivegauche}{result['url']}"
                    self.links_products.append(result_url)
            else:
                break

    def write_file_task_first(self) -> None:
        """
        Записывает информацию о продуктах в файл Excel во второй лист с названием "Задание 1" для первого задания.
        """
        workbook = load_workbook(filename=self.data_file)
        sheet = workbook['Задание 1']
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=7):
            link_cell = row[0]
            if link_cell.value in self.info_products:
                row[1].value = self.info_products[link_cell.value]['Цена до скидки']
                row[2].value = self.info_products[link_cell.value]['Цена со скидкой или по карте лояльности']
                row[3].value = self.info_products[link_cell.value]['Доступен для заказа (есть остаток)']
                row[4].value = self.info_products[link_cell.value]['Дата']
        workbook.save(filename='final_data_task_first.xlsx')

    def write_file(self, final_file_name: str, catalog: bool, brand: bool) -> None:
        """
        Записывает информацию о продуктах в файл Excel.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        headers = ['Наименование', 'Ссылка', 'Доступен для заказа (есть остаток)', 'Цена до скидки',
                   'Цена со скидкой или по карте лояльности', 'Дата']
        # опционально добавляется либо цепочка каталогов, либо бренд
        if catalog:
            headers.append('Цепочка категорий')
        elif brand:
            headers.append('Бренд')

        sheet.append(headers)

        for link, details in self.info_products.items():
            row = [details.get(header, '') for header in headers]
            sheet.append(row)

        workbook.save(final_file_name)

    def run_task_first(self) -> None:
        """
        Выполняет первую задачу: парсинг ссылок из файла и запись данных в Excel.
        """
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        self.get_links_from_file()
        asyncio.run(self.create_async_tasks())
        self.write_file_task_first()

    def run_task_second(self) -> None:
        """
        Выполняет вторую задачу: парсинг продуктов бренда Payot и запись данных в Excel.
        """
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        self.get_link_from_products(brand_code=self.brand_code_payot, category_code=None)
        asyncio.run(self.create_async_tasks())
        self.write_file(final_file_name='final_data_task_second.xlsx', catalog=True, brand=False)

    def run_task_third(self) -> None:
        """
        Выполняет третью задачу: парсинг продуктов категории кондиционеры и запись данных в Excel.
        """
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        self.get_link_from_products(brand_code=None, category_code=self.category_code_condtioner)
        asyncio.run(self.create_async_tasks())
        self.write_file(final_file_name='final_data_task_third.xlsx', catalog=False, brand=True)


if __name__ == '__main__':
    pars = ParserRivegauche()
    pars.run_task_first()
    pars.run_task_second()
    pars.run_task_third()
